# tests/test_admin.py
from django.test import TestCase, Client
from django.urls import reverse
from pc.models import User, Component
import openpyxl
from io import BytesIO


class AdminExportTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.admin = User.objects.create_user(
            username="admin", password="pass", role="admin", is_staff=True
        )
        Component.objects.create(
    name="RTX 4070",
    category="GPU",
    price=65000,
    has_pcie=True 
)

    def test_export_components_xlsx(self):
        """Экспорт компонентов в XLSX с правильными данными"""
        self.client.login(username="admin", password="pass")
        response = self.client.post(reverse('admin_export_xlsx'), {'model': 'Component'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn('filename="Component_export.xlsx"', response['Content-Disposition'])

        # Проверяем содержимое XLSX
        wb = openpyxl.load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.cell(1, 2).value, "Название")
        self.assertEqual(ws.cell(2, 2).value, "RTX 4070")
        self.assertEqual(ws.cell(2, 3).value, "GPU")
        self.assertEqual(ws.cell(2, 7).value, "Да")